"""
compare_sheets.py
-----------------
Compara dois ficheiros Excel (.xlsx) sheet a sheet, célula a célula.
Exporta as diferenças para um ficheiro Excel com highlighting.

Uso:
    python compare_sheets.py ficheiro_a.xlsx ficheiro_b.xlsx

Output:
    differences.xlsx  — só as linhas que diferem, com as duas versões lado a lado
    (se não houver diferenças, diz isso no terminal e não cria ficheiro)

Setup (só uma vez):
    pip install openpyxl
"""

import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.cell import MergedCell

# Cores para o output
FILL_A    = PatternFill("solid", fgColor="FFE0E0")  # vermelho claro — valor do ficheiro A
FILL_B    = PatternFill("solid", fgColor="E0FFE0")  # verde claro   — valor do ficheiro B
FILL_HEAD = PatternFill("solid", fgColor="D0D0D0")  # cinzento      — cabeçalho

def cell_value(cell):
    """Retorna o valor da célula como string normalizada para comparação."""
    if isinstance(cell, MergedCell):
        return ""
    if cell.value is None:
        return ""
    return str(cell.value).strip()

def compare_sheets(ws_a, ws_b):
    """
    Compara duas sheets célula a célula.
    Retorna lista de (row, col, value_a, value_b) para cada célula diferente.
    """
    max_row = max(ws_a.max_row, ws_b.max_row)
    max_col = max(ws_a.max_column, ws_b.max_column)
    diffs = []

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val_a = cell_value(ws_a.cell(row=r, column=c))
            val_b = cell_value(ws_b.cell(row=r, column=c))
            if val_a != val_b:
                diffs.append((r, c, val_a, val_b))

    return diffs

def col_letter(n):
    """Converte índice de coluna (1-based) para letra(s): 1→A, 27→AA."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

def main():
    if len(sys.argv) < 3:
        print("Uso: python compare_sheets.py ficheiro_a.xlsx ficheiro_b.xlsx")
        sys.exit(1)

    path_a = sys.argv[1]
    path_b = sys.argv[2]

    for p in (path_a, path_b):
        if not os.path.exists(p):
            print(f"ERRO: Ficheiro não encontrado: {p}")
            sys.exit(1)

    print(f"\nA carregar:")
    print(f"  A: {os.path.basename(path_a)}")
    print(f"  B: {os.path.basename(path_b)}")

    wb_a = load_workbook(path_a, data_only=True)
    wb_b = load_workbook(path_b, data_only=True)

    # Comparar sheets com o mesmo nome
    sheets_a = set(wb_a.sheetnames)
    sheets_b = set(wb_b.sheetnames)
    common   = sheets_a & sheets_b
    only_a   = sheets_a - sheets_b
    only_b   = sheets_b - sheets_a

    if only_a:
        print(f"\n  Sheets só em A: {', '.join(only_a)}")
    if only_b:
        print(f"  Sheets só em B: {', '.join(only_b)}")

    all_diffs = {}
    for sheet_name in sorted(common):
        print(f"\nA comparar sheet: '{sheet_name}'...")
        diffs = compare_sheets(wb_a[sheet_name], wb_b[sheet_name])
        all_diffs[sheet_name] = diffs
        print(f"  {len(diffs)} diferença(s) encontrada(s)")

    total = sum(len(d) for d in all_diffs.values())

    if total == 0:
        print("\n✓ Ficheiros idênticos — sem diferenças.")
        return

    # Criar ficheiro de output com as diferenças
    out_wb = Workbook()
    out_wb.remove(out_wb.active)  # remove sheet vazia default

    for sheet_name, diffs in all_diffs.items():
        if not diffs:
            continue

        ws_out = out_wb.create_sheet(title=sheet_name[:31])

        # Cabeçalho
        headers = ["Cell", "Row", "Col", f"A: {os.path.basename(path_a)}", f"B: {os.path.basename(path_b)}"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws_out.cell(row=1, column=c_idx, value=h)
            cell.fill = FILL_HEAD
            cell.font = Font(bold=True)

        # Largura das colunas
        ws_out.column_dimensions["A"].width = 8
        ws_out.column_dimensions["B"].width = 6
        ws_out.column_dimensions["C"].width = 6
        ws_out.column_dimensions["D"].width = 40
        ws_out.column_dimensions["E"].width = 40

        # Linhas de diferença
        for i, (row, col, val_a, val_b) in enumerate(diffs, start=2):
            cell_ref = f"{col_letter(col)}{row}"
            ws_out.cell(row=i, column=1, value=cell_ref)
            ws_out.cell(row=i, column=2, value=row)
            ws_out.cell(row=i, column=3, value=col)

            cell_a = ws_out.cell(row=i, column=4, value=val_a)
            cell_a.fill = FILL_A

            cell_b = ws_out.cell(row=i, column=5, value=val_b)
            cell_b.fill = FILL_B

    out_path = "differences.xlsx"
    out_wb.save(out_path)

    print(f"\n  Total de diferenças: {total}")
    print(f"  Output guardado em: {out_path}")
    print(f"\n  Legenda:")
    print(f"  🔴 Vermelho claro = valor em A ({os.path.basename(path_a)})")
    print(f"  🟢 Verde claro    = valor em B ({os.path.basename(path_b)})")

if __name__ == "__main__":
    main()
