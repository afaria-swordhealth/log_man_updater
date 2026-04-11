"""
update_logistics.py
===================
Reads a DHL invoice PDF and updates the Logistics Management Excel workbook
stored in Google Drive (synced locally via the Google Drive desktop app).

Usage
-----
    python update_logistics.py <path_to_invoice.pdf>

Example
-------
    python update_logistics.py "C:\\Users\\Andre\\Downloads\\LISR000538716.pdf"

Dependencies (install once)
---------------------------
    pip install pdfplumber openpyxl

Concurrency safety
------------------
The script records the file's last-modified timestamp before loading it.
Immediately before saving, it checks whether the timestamp has changed.
If the file was modified externally (e.g. someone edited it in Google Sheets
online and the Drive app synced it to disk), the script aborts and preserves
both versions: the original is kept and a timestamped backup of the in-memory
result is written alongside it for manual review.
"""

import sys
import re
import os
import shutil
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.worksheet.hyperlink import Hyperlink

# ── Configuration ─────────────────────────────────────────────────────────────

EXCEL_PATH = r"G:\My Drive\Logistics Management '26.xlsx"
SHEET_NAME = "LogMan 2026"

# Column indices (1-based) in the target sheet
COL_R = 18   # Shipping Document / Tracking Number
COL_U = 21   # Invoice Number
COL_V = 22   # Invoice Date
COL_W = 23   # Shipment Amount (EUR)


# ── Cell utilities ────────────────────────────────────────────────────────────

def normalize_tracking(val):
    """
    Normalise a tracking number value to a plain string.

    openpyxl reads numeric-looking cells as floats, so '3248136404'
    may be stored as 3248136404.0. This strips the trailing '.0' so
    that lookups match regardless of how the cell was originally typed.
    """
    if val is None:
        return None
    s = str(val).strip()
    return s[:-2] if re.match(r"^\d+\.0$", s) else s


def cell_is_red(cell):
    """
    Return True if the cell's font colour is red (R > 150, G < 100, B < 100).

    Handles both 6-digit (#RRGGBB) and 8-digit (#AARRGGBB) RGB strings.
    Silently returns False for theme colours or malformed values.
    """
    try:
        if cell.font and cell.font.color and cell.font.color.type == "rgb":
            rgb = cell.font.color.rgb.upper()
            if len(rgb) == 8:
                rgb = rgb[2:]          # strip alpha channel
            if len(rgb) == 6:
                r = int(rgb[0:2], 16)
                g = int(rgb[2:4], 16)
                b = int(rgb[4:6], 16)
                return r > 150 and g < 100 and b < 100
    except Exception:
        pass
    return False


def make_black_font(cell):
    """
    Return a copy of the cell's current font with the colour set to black.
    All other font attributes (name, size, bold, italic, etc.) are preserved.
    """
    f = cell.font
    return Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        underline=f.underline, strike=f.strike, color="FF000000",
    )


def clear_cell(cell):
    """
    Reset a cell's value and all formatting to defaults.
    MergedCell slaves are read-only in openpyxl and are silently skipped.
    """
    if isinstance(cell, MergedCell):
        return
    cell.value         = None
    cell.font          = Font()
    cell.fill          = PatternFill()
    cell.border        = Border()
    cell.alignment     = Alignment()
    cell.number_format = "General"
    cell.hyperlink     = None


# ── Sheet I/O helpers ─────────────────────────────────────────────────────────

def capture_row(ws, row_num, max_col):
    """
    Read one row from the sheet into a list of dicts.

    Each dict contains the cell's value and a deep copy of every style
    attribute so the data can be reconstructed later without the sheet.
    """
    row = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=c)
        row.append({
            "value":          cell.value,
            "font":           copy(cell.font)      if cell.has_style else None,
            "fill":           copy(cell.fill)      if cell.has_style else None,
            "border":         copy(cell.border)    if cell.has_style else None,
            "alignment":      copy(cell.alignment) if cell.has_style else None,
            "number_format":  cell.number_format,
            "hyperlink":      copy(cell.hyperlink) if cell.hyperlink else None,
        })
    return row


def write_row(ws, row_num, row_data):
    """
    Write a list of cell dicts (produced by capture_row or build_inserted_row)
    back into the sheet. MergedCell slaves are silently skipped.
    """
    for c_idx, cd in enumerate(row_data):
        cell = ws.cell(row=row_num, column=c_idx + 1)
        if isinstance(cell, MergedCell):
            continue
        cell.value = cd["value"]
        if cd["font"]:          cell.font          = cd["font"]
        if cd["fill"]:          cell.fill          = cd["fill"]
        if cd["border"]:        cell.border        = cd["border"]
        if cd["alignment"]:     cell.alignment     = cd["alignment"]
        if cd["number_format"]: cell.number_format = cd["number_format"]
        if cd.get("hyperlink"):   cell.hyperlink     = cd["hyperlink"]


def build_inserted_row(source_row, tracking, invoice_no, invoice_date, amount, max_col):
    """
    Build the cell-dict list for a new row inserted below an existing entry.

    Formatting is inherited from source_row (the row above).
    Columns U, V and W receive black text regardless of the source colour.
    Column R inherits the source formatting unchanged (per spec).
    All other columns are left empty but keep the source row's formatting.
    """
    new_row = []
    for c in range(1, max_col + 1):
        src  = source_row[c - 1]
        sf   = src["font"]

        # Determine value for this column
        if   c == COL_R: val = tracking
        elif c == COL_U: val = invoice_no
        elif c == COL_V: val = invoice_date
        elif c == COL_W: val = amount
        else:            val = None

        # U / V / W always use black text; R keeps the source font as-is
        if c in (COL_U, COL_V, COL_W):
            font = Font(
                name     = sf.name      if sf else None,
                size     = sf.size      if sf else None,
                bold     = sf.bold      if sf else None,
                italic   = sf.italic    if sf else None,
                underline= sf.underline if sf else None,
                strike   = sf.strike    if sf else None,
                color    = "FF000000",
            )
        else:
            font = copy(sf) if sf else None

        new_row.append({
            "value":         val,
            "font":          font,
            "fill":          copy(src["fill"])      if src["fill"]      else None,
            "border":        copy(src["border"])    if src["border"]    else None,
            "alignment":     copy(src["alignment"]) if src["alignment"] else None,
            "number_format": src["number_format"],
            "hyperlink":     None,   # new inserted row inherits no hyperlinks
        })
    return new_row


# ── PDF parsing ───────────────────────────────────────────────────────────────

def parse_invoice_pdf(pdf_path):
    """
    Extract invoice metadata and per-shipment amounts from a DHL invoice PDF.

    Page 1  → invoice number and invoice date (identical for every shipment).
    Page 2+ → one shipment block per tracking number.
              Each block begins with a 10-digit tracking number followed by
              a date (DD-MM-YYYY) and ends with a standalone decimal total
              (e.g. '126,09') on its own line.

    Returns
    -------
    invoice_no   : str   e.g. 'FT V/424543'
    invoice_date : datetime
    shipments    : dict  { tracking_number_str: amount_float }
    """
    import pdfplumber

    invoice_no       = None
    invoice_date     = None
    shipments        = {}
    current_tracking = None
    current_amount   = None

    with pdfplumber.open(pdf_path) as pdf:

        # ── Page 1: invoice header ────────────────────────────────────────────
        header_text = pdf.pages[0].extract_text() or ""
        for line in header_text.splitlines():
            line = line.strip()
            m = re.search(r"N[uú]mero da Fatura[:\s]+(FT\s*\S+)", line, re.IGNORECASE)
            if m:
                invoice_no = m.group(1).strip()
            m = re.search(r"Data Da Fatura[:\s]+(\d{2}-\d{2}-\d{4})", line, re.IGNORECASE)
            if m:
                invoice_date = datetime.strptime(m.group(1), "%d-%m-%Y")

        # ── Pages 2+: shipment detail lines ──────────────────────────────────
        for page in pdf.pages[1:]:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue

                # A new shipment block starts with a 10-digit tracking number.
                # The optional reference field may contain multiple tokens
                # (e.g. 'RTO 7981361555'), so we match the 10-digit prefix and
                # then verify a DD-MM-YYYY date appears anywhere later on the line.
                if re.match(r"^\d{10}\b", line) and re.search(r"\b\d{2}-\d{2}-\d{4}\b", line):
                    # Commit the previous shipment before starting the new one
                    if current_tracking and current_amount is not None:
                        shipments[current_tracking] = current_amount
                    current_tracking = line.split()[0]
                    current_amount   = None
                else:
                    # The shipment total appears as a standalone value on its own line.
                    # Handles plain decimals ('26,61') and amounts with a Portuguese
                    # thousands separator ('2.840,95', '5.572,24').
                    m = re.match(r"^(\d{1,3}(?:\.\d{3})*,\d{2}|\d+\.\d{2}|\d+,\d{2})$", line)
                    if m:
                        raw = m.group(1)
                        if "," in raw:
                            # PT format: period = thousands sep, comma = decimal sep
                            current_amount = float(raw.replace(".", "").replace(",", "."))
                        else:
                            current_amount = float(raw)

        # Commit the final shipment
        if current_tracking and current_amount is not None:
            shipments[current_tracking] = current_amount

    return invoice_no, invoice_date, shipments


# ── Concurrency check ─────────────────────────────────────────────────────────

def get_mtime(path):
    """Return the file's last-modified timestamp as a float."""
    return os.path.getmtime(path)


def save_with_conflict_check(wb, excel_path, mtime_before):
    """
    Save the workbook only if the source file has not been modified externally
    since it was loaded.

    If a modification is detected (Drive synced a newer version to disk while
    the script was running), the save is aborted. A timestamped backup of the
    in-memory result is written to the same directory so no work is lost.

    Returns True on a clean save, False on conflict.
    """
    mtime_now = get_mtime(excel_path)
    if mtime_now != mtime_before:
        # The file changed on disk — another save would overwrite those changes.
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = os.path.dirname(excel_path)
        backup_name = f"CONFLICT_BACKUP_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_name)
        wb.save(backup_path)
        print()
        print("  !! CONFLICT DETECTED !!")
        print(f"  The file was modified externally while the script was running.")
        print(f"  To prevent data loss the script did NOT overwrite the file.")
        print(f"  Your processed changes have been saved to:")
        print(f"  {backup_path}")
        print(f"  Please merge the two files manually.")
        return False

    wb.save(excel_path)
    return True


# ── Core update logic ─────────────────────────────────────────────────────────

def update_sheet(ws, invoice_no, invoice_date, shipments):
    """
    Apply invoice data from *shipments* to an already-loaded worksheet *ws*.

    Parameters
    ----------
    ws           : openpyxl Worksheet (already open, caller saves)
    invoice_no   : str   e.g. 'FT V/424543'
    invoice_date : datetime
    shipments    : dict  { tracking_number_str: amount_float }

    Returns
    -------
    (filled, inserted, created) : tuple[int, int, int]
        Number of rows filled in place, inserted below, and appended at end.
    """

    # Ensure max_col covers at least up to the last column we write to (COL_W)
    max_col = max(ws.max_column, COL_W)

    # Identify the last row that contains real data by scanning column A
    # downwards. This ignores stray values in other columns (e.g. orphaned
    # formula results) that can inflate ws.max_row far beyond the actual data.
    last_data_row = 1
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=1).value is not None:
            last_data_row = r
            break

    # Build a lookup table: normalised tracking number → row index.
    # IMPORTANT: scan the full sheet (not just up to last_data_row) because
    # rows appended by a previous run only have columns R/U/V/W filled —
    # column A is empty, so last_data_row stops before them. Without this
    # extended scan those rows would never be found and would be re-appended
    # on every subsequent run.
    existing_rows = {}
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=COL_R).value
        if val and str(val).strip() not in ("(Tracking link)", ""):
            existing_rows[normalize_tracking(val)] = r

    # ── Classify each shipment ────────────────────────────────────────────────

    # Group A: tracking exists in the sheet, column U is empty
    #          → fill U / V / W in place and remove any red formatting
    to_fill = []

    # Group B: tracking exists in the sheet, column U is already filled
    #          → the shipment spans multiple invoices; insert a new row below
    to_insert = []

    # Group C: tracking not found in the sheet at all
    #          → append a new row at the bottom of the data
    to_create = []

    skipped = 0
    for tracking, amount in shipments.items():
        key = normalize_tracking(tracking)
        if key in existing_rows:
            row_idx = existing_rows[key]
            existing_invoice = ws.cell(row=row_idx, column=COL_U).value

            if not existing_invoice:
                # Column U is empty — fill in place
                to_fill.append((tracking, amount, row_idx))

            elif str(existing_invoice).strip() == invoice_no.strip():
                # This exact invoice is already recorded for this tracking number.
                # Running the script twice must be safe — skip silently.
                skipped += 1

            else:
                # A different invoice exists — this shipment spans multiple invoices.
                # Insert a new row immediately below the existing entry.
                to_insert.append((tracking, amount, row_idx))
        else:
            to_create.append((tracking, amount))

    # ── Apply changes ─────────────────────────────────────────────────────────

    # ── Group A: fill in place ────────────────────────────────────────────────
    # Simple cell writes with no structural changes — fast.
    for tracking, amount, row_idx in to_fill:
        ws.cell(row=row_idx, column=COL_U, value=invoice_no)
        ws.cell(row=row_idx, column=COL_V, value=invoice_date)
        ws.cell(row=row_idx, column=COL_V).number_format = "DD-MM-YYYY"
        ws.cell(row=row_idx, column=COL_W, value=amount)
        ws.cell(row=row_idx, column=COL_W).number_format = "\u20ac#,##0.00"
        # Convert any red-coloured cells in U / V / W to black.
        # Column R is intentionally excluded — its formatting is not touched.
        for col in (COL_U, COL_V, COL_W):
            cell = ws.cell(row=row_idx, column=col)
            if cell_is_red(cell):
                cell.font = make_black_font(cell)

    # ── Pre-compute last occupied row (used by both Group B and Group C) ────────
    # existing_rows was built by scanning the full worksheet (ws.max_row), so
    # it includes rows appended by previous runs that only have COL_R filled
    # (col A empty → invisible to last_data_row).  This is the true boundary.
    last_occupied_row = max(existing_rows.values()) if existing_rows else last_data_row

    # ── Group B: insert rows below existing entries ───────────────────────────
    # Strategy: read the entire data range into memory (including any appended
    # rows beyond last_data_row), rebuild the row list with the insertions in
    # the correct positions, clear the affected area, then write everything
    # back in a single pass.  Appended rows are included in the snapshot so
    # that the clear + rewrite does not accidentally destroy them.
    if to_insert:
        # Build a mapping: source_row_index → list of (tracking, amount) to
        # insert immediately below that row.  Using a list supports the edge
        # case where the same tracking number appears in multiple invoices and
        # both are processed in the same run.
        insert_after: dict[int, list] = {}
        for tracking, amount, row_idx in to_insert:
            insert_after.setdefault(row_idx, []).append((tracking, amount))

        # Read main data rows (col A present) plus any appended rows beyond.
        rows_snapshot = [capture_row(ws, r, max_col) for r in range(1, last_data_row + 1)]
        # Rows beyond last_data_row: col A empty, only COL_R (and maybe U/V/W) filled.
        # Capture them so they survive the clear-and-rewrite below.
        appended_snapshot = [
            capture_row(ws, r, max_col)
            for r in range(last_data_row + 1, last_occupied_row + 1)
            if ws.cell(row=r, column=COL_R).value is not None
        ]

        # Reconstruct the row list, inserting new rows where required, then
        # append the previously-appended rows at the end (they shift up by
        # len(to_insert) to make room for the newly inserted rows).
        rebuilt_rows = []
        for i, row in enumerate(rows_snapshot):
            rebuilt_rows.append(row)
            original_row_num = i + 1
            for tracking, amount in insert_after.get(original_row_num, []):
                rebuilt_rows.append(
                    build_inserted_row(row, tracking, invoice_no, invoice_date, amount, max_col)
                )
        rebuilt_rows.extend(appended_snapshot)

        # Clear the entire area that will be rewritten (value + formatting).
        total_rows = len(rebuilt_rows)
        for r in range(1, total_rows + 1):
            for c in range(1, max_col + 1):
                clear_cell(ws.cell(row=r, column=c))

        for i, row in enumerate(rebuilt_rows):
            write_row(ws, i + 1, row)

    # ── Group C: append new rows at the end of the data ──────────────────────
    # After Group B, appended rows have been shifted up by len(to_insert).
    # new_last_row accounts for both the main data range and any appended rows.
    new_last_row = last_occupied_row + len(to_insert)
    next_row     = new_last_row + 1
    for tracking, amount in to_create:
        ws.cell(row=next_row, column=COL_R, value=tracking)
        ws.cell(row=next_row, column=COL_U, value=invoice_no)
        ws.cell(row=next_row, column=COL_V, value=invoice_date)
        ws.cell(row=next_row, column=COL_V).number_format = "DD-MM-YYYY"
        ws.cell(row=next_row, column=COL_W, value=amount)
        ws.cell(row=next_row, column=COL_W).number_format = "\u20ac#,##0.00"
        next_row += 1

    return len(to_fill), len(to_insert), len(to_create), skipped


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python update_logistics.py <path_to_invoice.pdf>")
        sys.exit(1)

    pdf_path = sys.argv[1]

    if not os.path.exists(pdf_path):
        print(f"ERROR: PDF not found: {pdf_path}")
        sys.exit(1)

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Workbook not found: {EXCEL_PATH}")
        print("Check that the Google Drive desktop app is running and synced.")
        sys.exit(1)

    # ── Step 1: Parse the PDF ─────────────────────────────────────────────────
    print(f"\n[1/4] Reading PDF...")
    invoice_no, invoice_date, shipments = parse_invoice_pdf(pdf_path)

    if not invoice_no or not invoice_date or not shipments:
        print("ERROR: Could not extract invoice data from the PDF.")
        sys.exit(1)

    print(f"      Invoice : {invoice_no}")
    print(f"      Date    : {invoice_date.strftime('%d-%m-%Y')}")
    print(f"      Entries : {len(shipments)} shipments")

    # ── Step 2: Load the workbook ─────────────────────────────────────────────
    print(f"\n[2/4] Loading workbook...")

    # Record the file timestamp BEFORE loading — used later for conflict detection
    mtime_before = get_mtime(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # ── Steps 3 & 4: Classify and apply ──────────────────────────────────────
    print(f"\n[3/4] Classifying shipments...")
    filled, inserted, created, skipped = update_sheet(ws, invoice_no, invoice_date, shipments)

    print(f"      Fill in place  : {filled}")
    print(f"      Insert below   : {inserted}")
    print(f"      Append at end  : {created}")
    if skipped:
        print(f"      Already done   : {skipped}  ← invoice already recorded, no changes needed")

    print(f"\n[4/4] Applying changes...")

    # ── Save with conflict detection ──────────────────────────────────────────
    saved = save_with_conflict_check(wb, EXCEL_PATH, mtime_before)

    if saved:
        print(f"\n  Done.")
        print(f"  Filled in place : {filled}")
        print(f"  Inserted below  : {inserted}")
        print(f"  Appended at end : {created}")
        print(f"\n  Google Drive will sync the updated file automatically.")


if __name__ == "__main__":
    main()
