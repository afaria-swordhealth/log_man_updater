"""
tests/test_integration.py
=========================
Integration tests for update_logistics.update_sheet().

All tests build in-memory workbooks using openpyxl directly — no PDFs or
external fixtures are required.  Each test saves to a NamedTemporaryFile,
reloads it (exactly as the real main() does), calls update_sheet(), and
then asserts against the resulting worksheet state.
"""

import os
import tempfile
from copy import copy
from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# Module under test
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from update_logistics import update_sheet, COL_R, COL_U, COL_V, COL_W


# ── Helpers ───────────────────────────────────────────────────────────────────

SHEET_NAME = "LogMan 2026"

INVOICE_NO   = "FT V/000099"
INVOICE_DATE = datetime(2026, 4, 11)
AMOUNT       = 126.09


def _make_wb(rows_spec):
    """
    Build a temporary .xlsx file from a list of row specification dicts.

    Row index starts at 2 (row 1 is always the header "Header").
    Each dict may contain:
        col_a     : value for column A  (marks a 'real' data row)
        tracking  : value for column R
        invoice   : value for column U
        date      : value for column V
        amount    : value for column W
        red_uvw   : if True, set a red font on U/V/W cells

    Returns the path to the saved temporary file.
    """
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.cell(row=1, column=1, value="Header")

    red_font  = Font(color="FFFF0000")
    black_font = Font(color="FF000000")

    for i, spec in enumerate(rows_spec, start=2):
        if spec.get("col_a") is not None:
            ws.cell(row=i, column=1, value=spec["col_a"])
        if spec.get("tracking") is not None:
            ws.cell(row=i, column=COL_R, value=spec["tracking"])
        if spec.get("invoice") is not None:
            ws.cell(row=i, column=COL_U, value=spec["invoice"])
        if spec.get("date") is not None:
            ws.cell(row=i, column=COL_V, value=spec["date"])
        if spec.get("amount") is not None:
            ws.cell(row=i, column=COL_W, value=spec["amount"])
        if spec.get("red_uvw"):
            for col in (COL_U, COL_V, COL_W):
                ws.cell(row=i, column=col).font = copy(red_font)

    wb.save(tmp.name)
    return tmp.name


def _run(path, invoice_no=INVOICE_NO, invoice_date=INVOICE_DATE, shipments=None):
    """
    Load *path*, run update_sheet, save, reload and return the worksheet.
    This mirrors exactly what main() does (minus PDF parsing and conflict check).
    """
    if shipments is None:
        shipments = {"1234567890": AMOUNT}

    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    filled, inserted, created, skipped = update_sheet(ws, invoice_no, invoice_date, shipments)
    wb.save(path)

    wb2 = load_workbook(path)
    ws2 = wb2[SHEET_NAME]
    return ws2, filled, inserted, created


# ── Tests ─────────────────────────────────────────────────────────────────────

class TestGroupA:

    def test_fill_in_place(self, tmp_path):
        """Tracking exists with empty U/V/W → cells are filled correctly."""
        path = _make_wb([
            {"col_a": "Ampere", "tracking": "1234567890"},
        ])
        try:
            ws, filled, inserted, created = _run(path)

            assert filled  == 1
            assert inserted == 0
            assert created  == 0

            assert ws.cell(row=2, column=COL_U).value == INVOICE_NO
            assert ws.cell(row=2, column=COL_V).value == INVOICE_DATE
            assert ws.cell(row=2, column=COL_W).value == AMOUNT
            assert ws.cell(row=2, column=COL_W).number_format == "\u20ac#,##0.00"
        finally:
            os.unlink(path)

    def test_removes_red_font(self, tmp_path):
        """Bug 11/12 regression: red font on U/V/W must be converted to black."""
        path = _make_wb([
            {"col_a": "Ampere", "tracking": "1234567890",
             "invoice": None, "red_uvw": True},
        ])
        try:
            ws, *_ = _run(path)

            for col in (COL_U, COL_V, COL_W):
                cell = ws.cell(row=2, column=col)
                font = cell.font
                # Font colour must NOT be red after update
                if font and font.color and font.color.type == "rgb":
                    rgb = font.color.rgb.upper()
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    r_val = int(rgb[0:2], 16)
                    g_val = int(rgb[2:4], 16)
                    b_val = int(rgb[4:6], 16)
                    assert not (r_val > 150 and g_val < 100 and b_val < 100), \
                        f"Column {col} still has red font after update"
        finally:
            os.unlink(path)

    def test_idempotent(self, tmp_path):
        """Bug 7 regression: running twice with same invoice → second run is a no-op."""
        path = _make_wb([
            {"col_a": "Ampere", "tracking": "1234567890"},
        ])
        try:
            # First run
            _run(path)
            # Second run
            ws, filled, inserted, created = _run(path)

            assert filled  == 0
            assert inserted == 0
            assert created  == 0

            # U still has the invoice from the first run, not a duplicate
            assert ws.cell(row=2, column=COL_U).value == INVOICE_NO
            # No extra rows were appended
            tracking_values = [
                ws.cell(row=r, column=COL_R).value
                for r in range(2, ws.max_row + 1)
            ]
            assert tracking_values.count("1234567890") == 1
        finally:
            os.unlink(path)


class TestGroupB:

    def test_insert_below(self, tmp_path):
        """Tracking exists with a different invoice in U → new row inserted below."""
        path = _make_wb([
            {"col_a": "Ampere", "tracking": "1234567890",
             "invoice": "FT V/000001", "date": datetime(2026, 1, 10), "amount": 55.0},
        ])
        try:
            ws, filled, inserted, created = _run(path)

            assert filled  == 0
            assert inserted == 1
            assert created  == 0

            # Original row 2 must be untouched
            assert ws.cell(row=2, column=COL_U).value == "FT V/000001"
            assert ws.cell(row=2, column=COL_R).value == "1234567890"

            # New row below must carry the new invoice
            assert ws.cell(row=3, column=COL_U).value == INVOICE_NO
            assert ws.cell(row=3, column=COL_R).value == "1234567890"
        finally:
            os.unlink(path)


class TestGroupC:

    def test_append_end(self, tmp_path):
        """Tracking not in sheet → new row created at the end."""
        path = _make_wb([
            {"col_a": f"Row{i}", "tracking": f"99999{i:05d}"} for i in range(1, 6)
        ])
        try:
            ws, filled, inserted, created = _run(
                path, shipments={"1234567890": AMOUNT}
            )

            assert created == 1

            # Row 7 = header(1) + 5 data rows + new row
            new_row = 7
            assert ws.cell(row=new_row, column=COL_R).value == "1234567890"
            assert ws.cell(row=new_row, column=COL_U).value == INVOICE_NO
            assert ws.cell(row=new_row, column=COL_V).value == INVOICE_DATE
            assert ws.cell(row=new_row, column=COL_W).value == AMOUNT
        finally:
            os.unlink(path)

    def test_no_overwrite_appended_row(self, tmp_path):
        """
        Bug 9 regression — THE MAIN BUG.

        Workbook state:
          Rows 2-5 : normal rows (col A + tracking in R, U/V/W filled)
          Row 6    : 'Ampere' row — ONLY tracking in R, col A empty, U/V/W empty

        A shipment with a BRAND NEW tracking must NOT overwrite row 6.
        It must be appended to row 7.
        """
        existing_tracking = "9000000001"
        ampere_tracking   = "8888888888"
        new_tracking      = "7777777777"

        rows_spec = [
            {"col_a": f"Row{i}", "tracking": f"900000000{i}",
             "invoice": "FT V/000001", "date": datetime(2026, 1, 1), "amount": 10.0}
            for i in range(1, 5)
        ]
        # Row 6: appended row — only R filled, col A empty
        rows_spec.append({"tracking": ampere_tracking})

        path = _make_wb(rows_spec)
        try:
            ws, filled, inserted, created = _run(
                path, shipments={new_tracking: AMOUNT}
            )

            assert created == 1

            # Row 6 (Ampere's row) must be intact
            assert ws.cell(row=6, column=COL_R).value == ampere_tracking
            assert ws.cell(row=6, column=COL_U).value is None
            assert ws.cell(row=6, column=1).value is None  # col A still empty

            # New tracking must be on row 7
            assert ws.cell(row=7, column=COL_R).value == new_tracking
            assert ws.cell(row=7, column=COL_U).value == INVOICE_NO
        finally:
            os.unlink(path)


class TestBugRegressions:

    def test_orphan_cells_dont_affect_last_data_row(self, tmp_path):
        """
        Bug 1 regression: a stray value in column Z row 500 must not shift
        last_data_row beyond the real data area.  New appended rows must land
        immediately after the last col-A row, not at row 501.
        """
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        path = tmp.name

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.cell(row=1, column=1, value="Header")
        for i in range(1, 11):
            ws.cell(row=i + 1, column=1, value=f"Row{i}")
            ws.cell(row=i + 1, column=COL_R, value=f"1000000{i:03d}")
        # Orphan cell far beyond data
        ws.cell(row=500, column=26, value="orphan")
        wb.save(path)

        try:
            new_tracking = "5555555555"
            ws2, filled, inserted, created = _run(
                path, shipments={new_tracking: AMOUNT}
            )

            assert created == 1

            # New row must be row 12 (header=1, data rows 2-11, new=12)
            assert ws2.cell(row=12, column=COL_R).value == new_tracking
            # Must NOT be at row 501
            assert ws2.cell(row=501, column=COL_R).value is None
        finally:
            os.unlink(path)

    def test_float_tracking_lookup(self, tmp_path):
        """
        Bug 2 regression: tracking stored as integer in the cell must still
        match a string tracking number from the PDF.
        """
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        path = tmp.name

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="SomeRow")
        ws.cell(row=2, column=COL_R, value=3248136404)   # stored as int
        wb.save(path)

        try:
            # PDF would give us the string "3248136404"
            ws2, filled, inserted, created = _run(
                path, shipments={"3248136404": AMOUNT}
            )

            # Must match (Group A) — NOT fall through to Group C
            assert filled  == 1
            assert created == 0
            assert ws2.cell(row=2, column=COL_U).value == INVOICE_NO
        finally:
            os.unlink(path)

    def test_multiple_appended_rows_preserved(self, tmp_path):
        """
        Bug 9 extended: 5 normal rows + 3 appended rows already present.
        Adding 2 new shipments must place them after the 3 appended rows and
        leave all 3 appended rows intact.
        """
        rows_spec = [
            {"col_a": f"Row{i}", "tracking": f"1100000{i:03d}",
             "invoice": "FT V/000001", "date": datetime(2026, 1, 1), "amount": 10.0}
            for i in range(1, 6)
        ]
        appended_trackings = ["2200000001", "2200000002", "2200000003"]
        for t in appended_trackings:
            rows_spec.append({"tracking": t})   # col A empty

        path = _make_wb(rows_spec)
        try:
            new_shipments = {"3300000001": 50.0, "3300000002": 60.0}
            ws, filled, inserted, created = _run(
                path, shipments=new_shipments
            )

            assert created == 2

            # All 3 appended rows must still be present and unchanged
            all_r_values = [
                ws.cell(row=r, column=COL_R).value
                for r in range(2, ws.max_row + 1)
            ]
            for t in appended_trackings:
                assert t in all_r_values, f"Appended tracking {t} was lost"

            # New rows must exist too
            for t in new_shipments:
                assert t in all_r_values, f"New tracking {t} was not created"
        finally:
            os.unlink(path)

    def test_date_format(self, tmp_path):
        """Bug 12 regression: V cell number_format must be 'DD-MM-YYYY' after fill."""
        path = _make_wb([
            {"col_a": "Row1", "tracking": "1234567890"},
        ])
        try:
            ws, *_ = _run(path)
            assert ws.cell(row=2, column=COL_V).number_format == "DD-MM-YYYY"
        finally:
            os.unlink(path)

    def test_amount_format(self, tmp_path):
        """Bug 11 regression: W cell number_format must be '€#,##0.00' after fill."""
        path = _make_wb([
            {"col_a": "Row1", "tracking": "1234567890"},
        ])
        try:
            ws, *_ = _run(path)
            assert ws.cell(row=2, column=COL_W).number_format == "\u20ac#,##0.00"
        finally:
            os.unlink(path)

    def test_no_overwrite_row_with_data_but_no_tracking(self, tmp_path):
        """
        Bug 13 regression: rows beyond last_data_row may have real data in
        cols B-AH even when col R (tracking) is still empty — e.g. export rows
        that are pre-filled with metadata before the invoice arrives.

        A brand-new Group C append must land AFTER these rows, not on top of them.
        """
        import tempfile as _tmp
        tmp = _tmp.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        path = tmp.name

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.cell(row=1, column=1, value="Header")

        # 3 normal rows (col A + tracking + invoice)
        for i in range(1, 4):
            ws.cell(row=i + 1, column=1,      value=f"Row{i}")
            ws.cell(row=i + 1, column=COL_R,  value=f"1000000{i:03d}")
            ws.cell(row=i + 1, column=COL_U,  value="FT V/000001")
            ws.cell(row=i + 1, column=COL_V,  value=datetime(2026, 1, 1))
            ws.cell(row=i + 1, column=COL_W,  value=10.0)

        # Row 5: export row — col A empty, col R empty, but has metadata
        # in other columns (mimics real rows 2130-2131)
        ws.cell(row=5, column=2,   value="Supply Chain & Logistics")  # col B
        ws.cell(row=5, column=3,   value="Export")                    # col C
        ws.cell(row=5, column=4,   value="PT - US")                   # col D
        ws.cell(row=5, column=8,   value="K-Log")                     # col H
        ws.cell(row=5, column=19,  value="Booked")                    # col S

        wb.save(path)
        try:
            new_tracking = "7777777777"
            ws2, filled, inserted, created = _run(
                path, shipments={new_tracking: AMOUNT}
            )

            assert created == 1

            # Row 5 must be untouched — col B still has its original value
            assert ws2.cell(row=5, column=2).value == "Supply Chain & Logistics", \
                "Export metadata row (col A empty, col R empty) was overwritten"
            assert ws2.cell(row=5, column=COL_R).value is None, \
                "Col R of pre-existing export row was overwritten"

            # New tracking must land at row 6 (after the export metadata row)
            assert ws2.cell(row=6, column=COL_R).value == new_tracking, \
                f"New tracking landed at wrong row (expected row 6, got overwrite at row 5)"
        finally:
            os.unlink(path)

    def test_insert_preserves_appended_rows(self, tmp_path):
        """
        Bug 9 + Group B combined: a Group B insert must not destroy appended rows.
        Setup:
          - 5 normal rows (col A + invoice already filled in U)
          - 1 appended row (only R, col A empty)
        Actions:
          - 1 shipment with same tracking as row 2 but new invoice → Group B insert
          - 1 brand-new shipment → Group C append
        Expected:
          - Appended row survives
          - Group C row lands after the appended row
        """
        appended_tracking = "8800000001"
        existing_tracking = "1100000001"

        rows_spec = [
            {"col_a": f"Row{i}", "tracking": f"1100000{i:03d}",
             "invoice": "FT V/000001", "date": datetime(2026, 1, 1), "amount": 10.0}
            for i in range(1, 6)
        ]
        rows_spec.append({"tracking": appended_tracking})   # col A empty

        path = _make_wb(rows_spec)
        try:
            new_tracking = "9900000001"
            shipments = {
                existing_tracking: AMOUNT,   # Group B (U already has FT V/000001)
                new_tracking:       75.0,    # Group C
            }
            ws, filled, inserted, created = _run(path, shipments=shipments)

            assert inserted == 1
            assert created  == 1

            all_r_values = [
                ws.cell(row=r, column=COL_R).value
                for r in range(2, ws.max_row + 1)
            ]

            # Appended row must survive
            assert appended_tracking in all_r_values, \
                "Appended row was destroyed by Group B rebuild"

            # Group C row must exist
            assert new_tracking in all_r_values, \
                "Group C row was not created"

            # Group C row must come AFTER the appended row
            idx_appended = all_r_values.index(appended_tracking)
            idx_new      = all_r_values.index(new_tracking)
            assert idx_new > idx_appended, \
                "Group C row landed before the appended row"
        finally:
            os.unlink(path)
