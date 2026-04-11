"""
tests/test_unit.py
==================
Unit tests for the pure-logic helpers in update_logistics.py.

No I/O or PDF parsing is exercised here; all tests work on in-memory
openpyxl objects.
"""

import sys
import os
from copy import copy
from datetime import datetime

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.cell import MergedCell

# ---------------------------------------------------------------------------
# Make the repo root importable regardless of where pytest is invoked from
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from update_logistics import (
    normalize_tracking,
    cell_is_red,
    make_black_font,
    clear_cell,
    build_inserted_row,
    COL_R, COL_U, COL_V, COL_W,
)


# ===========================================================================
# Helpers
# ===========================================================================

def _make_wb():
    """Return a fresh Workbook with one active worksheet."""
    wb = Workbook()
    return wb, wb.active


def _cell(value=None, font=None, row=1, col=1):
    """Create a real openpyxl cell in a fresh workbook and return it."""
    _, ws = _make_wb()
    c = ws.cell(row=row, column=col, value=value)
    if font is not None:
        c.font = font
    return c


# ===========================================================================
# normalize_tracking
# ===========================================================================

class TestNormalizeTracking:

    def test_none_returns_none(self):
        assert normalize_tracking(None) is None

    def test_plain_string_unchanged(self):
        assert normalize_tracking("3248136404") == "3248136404"

    def test_float_strips_dot_zero(self):
        """Bug 2 regression: openpyxl reads numeric cells as float."""
        assert normalize_tracking(3248136404.0) == "3248136404"

    def test_string_with_dot_zero_stripped(self):
        """String variant of Bug 2 — e.g. value already converted to str."""
        assert normalize_tracking("3248136404.0") == "3248136404"

    def test_strip_whitespace(self):
        assert normalize_tracking("  3248136404  ") == "3248136404"

    def test_non_numeric_string_unchanged(self):
        assert normalize_tracking("ABC123") == "ABC123"

    def test_real_decimal_not_stripped(self):
        """A genuine decimal (non-integer) must NOT be modified."""
        assert normalize_tracking("12345.50") == "12345.50"

    def test_zero_float_edge_case(self):
        assert normalize_tracking(0.0) == "0"


# ===========================================================================
# cell_is_red
# ===========================================================================

class TestCellIsRed:

    def test_no_font_returns_false(self):
        _, ws = _make_wb()
        cell = ws.cell(row=1, column=1, value="x")
        # Default cell has no font colour set
        assert cell_is_red(cell) is False

    def test_black_font_returns_false(self):
        cell = _cell(font=Font(color="FF000000"))
        assert cell_is_red(cell) is False

    def test_red_six_digit_returns_true(self):
        """Pure red in 6-digit RGB format."""
        cell = _cell(font=Font(color="FF0000"))
        assert cell_is_red(cell) is True

    def test_red_eight_digit_returns_true(self):
        """Pure red with alpha prefix (8-digit ARGB)."""
        cell = _cell(font=Font(color="FFFF0000"))
        assert cell_is_red(cell) is True

    def test_borderline_red_above_threshold_true(self):
        """R=151, G=99, B=99 — just above the threshold, should be True."""
        # Encode as 8-digit ARGB: FF + hex(R) + hex(G) + hex(B)
        color = "FF{:02X}{:02X}{:02X}".format(151, 99, 99)
        cell = _cell(font=Font(color=color))
        assert cell_is_red(cell) is True

    def test_borderline_red_at_threshold_false(self):
        """R=150, G=100, B=100 — exactly at boundary, should be False (> not >=)."""
        color = "FF{:02X}{:02X}{:02X}".format(150, 100, 100)
        cell = _cell(font=Font(color=color))
        assert cell_is_red(cell) is False

    def test_theme_colour_returns_false_no_crash(self):
        """Theme colours have type != 'rgb'; must not crash and return False."""
        from openpyxl.styles.colors import Color
        _, ws = _make_wb()
        cell = ws.cell(row=1, column=1, value="x")
        # Build a theme-based colour directly
        theme_color = Color(theme=4)
        cell.font = Font(color=theme_color)
        assert cell_is_red(cell) is False


# ===========================================================================
# make_black_font
# ===========================================================================

class TestMakeBlackFont:

    def test_preserves_all_attributes(self):
        """All six non-colour attributes must survive the copy."""
        original = Font(
            name="Calibri",
            size=12,
            bold=True,
            italic=True,
            underline="single",
            strike=True,
            color="FFFF0000",   # red — should become black
        )
        cell = _cell(font=original)
        new_font = make_black_font(cell)

        assert new_font.name == "Calibri"
        assert new_font.size == 12
        assert new_font.bold is True
        assert new_font.italic is True
        assert new_font.underline == "single"
        assert new_font.strike is True

    def test_colour_becomes_black(self):
        cell = _cell(font=Font(color="FFFF0000"))
        new_font = make_black_font(cell)
        assert new_font.color.rgb.upper() == "FF000000"

    def test_none_attributes_do_not_crash(self):
        """Cell with a minimal/default Font must not raise."""
        _, ws = _make_wb()
        cell = ws.cell(row=1, column=1, value="x")
        # openpyxl default font — most attributes are None/False
        result = make_black_font(cell)
        assert result.color.rgb.upper() == "FF000000"


# ===========================================================================
# clear_cell
# ===========================================================================

class TestClearCell:

    def test_normal_cell_is_cleared(self):
        """Value, font, fill, border, alignment, number_format and hyperlink
        must all be reset to their defaults."""
        _, ws = _make_wb()
        cell = ws.cell(row=1, column=1, value="hello")
        cell.font = Font(color="FFFF0000", bold=True)
        cell.number_format = "0.00"

        clear_cell(cell)

        assert cell.value is None
        assert cell.font == Font()
        assert cell.fill == PatternFill()
        assert cell.border == Border()
        assert cell.alignment == Alignment()
        assert cell.number_format == "General"
        assert cell.hyperlink is None

    def test_merged_cell_slave_does_not_crash(self):
        """Bug 5 regression: MergedCell slaves are read-only; clear_cell
        must silently skip them without raising AttributeError."""
        _, ws = _make_wb()
        ws.merge_cells("A2:B2")
        merged_slave = ws.cell(row=2, column=2)   # column B is the slave

        assert isinstance(merged_slave, MergedCell), (
            "Test setup error: expected a MergedCell slave"
        )
        # Must not raise
        clear_cell(merged_slave)


# ===========================================================================
# build_inserted_row
# ===========================================================================

def _make_source_row(max_col, red=True):
    """
    Build a source_row list (as returned by capture_row) with:
    - Col R  (COL_R=18): tracking "ORIG_TRACK"
    - Col U  (COL_U=21): number_format '€#,##0.00' (Bug 11 format preset)
    - Col V  (COL_V=22): number_format 'DD-MM-YYYY'  (Bug 12 format preset)
    - Col W  (COL_W=23): number_format '€#,##0.00'
    - All cols: red font if red=True
    """
    _, ws = _make_wb()
    red_font = Font(
        name="Arial", size=10, bold=False, italic=False,
        underline=None, strike=False, color="FFFF0000",
    )
    row = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c, value="src" if c == COL_R else None)
        if red:
            cell.font = red_font

        # Preset number_format so that build_inserted_row can inherit it
        if c == COL_V:
            cell.number_format = "DD-MM-YYYY"
        elif c in (COL_U, COL_W):
            cell.number_format = "\u20AC#,##0.00"
        else:
            cell.number_format = "General"

        row.append({
            "value":         cell.value,
            "font":          copy(cell.font) if cell.has_style else None,
            "fill":          copy(cell.fill) if cell.has_style else None,
            "border":        copy(cell.border) if cell.has_style else None,
            "alignment":     copy(cell.alignment) if cell.has_style else None,
            "number_format": cell.number_format,
            "hyperlink":     copy(cell.hyperlink) if cell.hyperlink else None,
        })
    return row


class TestBuildInsertedRow:

    MAX_COL = 25  # wide enough to cover all named columns

    def _build(self, **kwargs):
        defaults = dict(
            tracking="9876543210",
            invoice_no="FT V/999",
            invoice_date=datetime(2026, 4, 1),
            amount=126.09,
            max_col=self.MAX_COL,
        )
        defaults.update(kwargs)
        source = _make_source_row(defaults["max_col"])
        return build_inserted_row(
            source,
            defaults["tracking"],
            defaults["invoice_no"],
            defaults["invoice_date"],
            defaults["amount"],
            defaults["max_col"],
        )

    # ── Length ─────────────────────────────────────────────────────────────

    def test_row_length_equals_max_col(self):
        row = self._build()
        assert len(row) == self.MAX_COL

    # ── Values in named columns ────────────────────────────────────────────

    def test_col_r_has_tracking(self):
        row = self._build(tracking="9876543210")
        assert row[COL_R - 1]["value"] == "9876543210"

    def test_col_u_has_invoice_no(self):
        row = self._build(invoice_no="FT V/999")
        assert row[COL_U - 1]["value"] == "FT V/999"

    def test_col_v_has_invoice_date(self):
        d = datetime(2026, 4, 1)
        row = self._build(invoice_date=d)
        assert row[COL_V - 1]["value"] == d

    def test_col_w_has_amount(self):
        row = self._build(amount=126.09)
        assert row[COL_W - 1]["value"] == pytest.approx(126.09)

    # ── number_format — Bug 11 & 12 regression ─────────────────────────────

    def test_col_v_number_format_date(self):
        """Bug 12: date cell must inherit 'DD-MM-YYYY' from the source row."""
        row = self._build()
        assert row[COL_V - 1]["number_format"] == "DD-MM-YYYY"

    def test_col_w_number_format_currency(self):
        """Bug 11: amount cell must inherit '€#,##0.00' from the source row."""
        row = self._build()
        assert row[COL_W - 1]["number_format"] == "\u20AC#,##0.00"

    def test_col_u_number_format_currency(self):
        """Invoice number cell inherits its number_format from the source row."""
        row = self._build()
        assert row[COL_U - 1]["number_format"] == "\u20AC#,##0.00"

    # ── Font colour for U / V / W (must be black, never red) ──────────────

    def test_col_u_font_is_black(self):
        row = self._build()
        font = row[COL_U - 1]["font"]
        assert font is not None
        assert font.color.rgb.upper() == "FF000000"

    def test_col_v_font_is_black(self):
        row = self._build()
        font = row[COL_V - 1]["font"]
        assert font is not None
        assert font.color.rgb.upper() == "FF000000"

    def test_col_w_font_is_black(self):
        row = self._build()
        font = row[COL_W - 1]["font"]
        assert font is not None
        assert font.color.rgb.upper() == "FF000000"

    # ── U / V / W preserve non-colour font attributes from source row ──────

    def test_col_u_preserves_font_name(self):
        row = self._build()
        assert row[COL_U - 1]["font"].name == "Arial"

    def test_col_v_preserves_font_size(self):
        row = self._build()
        assert row[COL_V - 1]["font"].size == 10

    # ── Other columns: value=None, formatting inherited ───────────────────

    def test_other_columns_value_is_none(self):
        row = self._build()
        other_cols = [
            c for c in range(1, self.MAX_COL + 1)
            if c not in (COL_R, COL_U, COL_V, COL_W)
        ]
        for c in other_cols:
            assert row[c - 1]["value"] is None, (
                f"Column {c} should have value=None but got {row[c - 1]['value']!r}"
            )

    def test_other_columns_inherit_formatting(self):
        """Columns outside R/U/V/W should carry the source row's font."""
        row = self._build()
        # Column 1 is not a special column; source has red font
        font = row[0]["font"]
        assert font is not None
        # Should inherit red from source (not forced to black)
        assert font.color.rgb.upper() == "FFFF0000"

    # ── hyperlink is never inherited ───────────────────────────────────────

    def test_all_hyperlinks_are_none(self):
        row = self._build()
        for i, cell_dict in enumerate(row):
            assert cell_dict["hyperlink"] is None, (
                f"Column {i + 1} should have hyperlink=None"
            )
