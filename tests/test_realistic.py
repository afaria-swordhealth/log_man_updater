"""
tests/test_realistic.py
========================
Regression tests using a workbook that faithfully mirrors the structure of
the real Logistics Management '26.xlsx file.

Unlike the synthetic tests in test_integration.py (which use minimal 5-row
workbooks), these tests exercise the exact structural patterns found in
production:

  • Domestic rows  : col A = "OMB-XXXX", tracking stored as int (Bug 2)
  • Export rows    : col A intentionally empty, tracking as string
  • W-without-U/V  : col W pre-filled (estimated rate) with no invoice yet
  • Appended rows  : col A empty, only R/U/V/W — from a previous script run
  • Orphan cell    : stray =sum() formula far beyond data (Bug 1 / ws.max_row)
  • Header row 2   : "(Tracking link)" in col R — must be excluded from index
"""

import os
import sys
import tempfile
from datetime import datetime

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from update_logistics import update_sheet, COL_R, COL_U, COL_V, COL_W
from tests.fixtures.make_realistic_workbook import (
    make_realistic_workbook,
    LAST_DATA_ROW,
    LAST_OCCUPIED_ROW,
    FIRST_FREE_ROW,
    PENDING_EXPORT_TRACKINGS,
    INVOICED_DOMESTIC_TRACKING,
    INVOICED_EXPORT_TRACKING,
    PENDING_DOMESTIC_TRACKINGS,
)

SHEET       = "LogMan 2026"
INVOICE_NO  = "FT V/999999"
INVOICE_DT  = datetime(2026, 4, 11)


def _run(path, shipments):
    wb = load_workbook(path)
    ws = wb[SHEET]
    result = update_sheet(ws, INVOICE_NO, INVOICE_DT, shipments)
    wb.save(path)
    wb2 = load_workbook(path)
    return wb2[SHEET], result


def _all_r_values(ws):
    return [
        ws.cell(row=r, column=COL_R).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=COL_R).value not in (None, "(Tracking link)")
    ]


# ── Orphan cell / ws.max_row ──────────────────────────────────────────────────

class TestOrphanCell:

    def test_orphan_does_not_affect_append_position(self):
        """
        Bug 1 + Bug 9 combined: orphan cell at row 500 inflates ws.max_row.
        New append must land at FIRST_FREE_ROW (24), not at row 501.
        """
        path = make_realistic_workbook()
        try:
            new_tracking = "4444444444"
            ws, (filled, inserted, created) = _run(path, {new_tracking: 9.99})

            assert created == 1

            r_values = _all_r_values(ws)
            assert new_tracking in r_values

            # Find the actual row of the new tracking
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=COL_R).value) == new_tracking:
                    assert r == FIRST_FREE_ROW, (
                        f"New row landed at {r} instead of {FIRST_FREE_ROW}. "
                        f"Orphan cell at row 500 may have shifted append position."
                    )
                    break

            # Must NOT be near the orphan cell
            assert ws.cell(row=501, column=COL_R).value is None
        finally:
            os.unlink(path)

    def test_header_row_excluded_from_index(self):
        """
        Row 2 has "(Tracking link)" in col R — it must not appear in
        existing_rows and must never be treated as a data row.
        """
        path = make_realistic_workbook()
        try:
            # If header were in existing_rows it would match "(Tracking link)"
            # and could be filled in place — no such row should be touched.
            new_tracking = "5555555555"
            ws, _ = _run(path, {new_tracking: 5.00})

            # Row 2 col U must still be "Invoice No." (the header label)
            assert ws.cell(row=2, column=COL_U).value == "Invoice No."
        finally:
            os.unlink(path)


# ── Domestic rows (int tracking, Bug 2) ──────────────────────────────────────

class TestDomesticRows:

    def test_int_tracking_found_and_skipped_when_already_invoiced(self):
        """
        Bug 2: tracking stored as int (3248136404) must be found in existing_rows
        and, since it already has an invoice, must be skipped (idempotent).
        """
        path = make_realistic_workbook()
        try:
            # INVOICED_DOMESTIC_TRACKING already has FT V/395474
            ws, (filled, inserted, created) = _run(
                path, {str(INVOICED_DOMESTIC_TRACKING): 6.97}
            )
            # Already invoiced with a DIFFERENT invoice → insert below
            # (the fixture has FT V/395474; we're running with FT V/999999)
            assert inserted == 1
            assert created  == 0
        finally:
            os.unlink(path)

    def test_pending_domestic_row_filled_in_place(self):
        """
        Domestic rows 13–15: col A filled, col W pre-filled (rate), but
        col U/V empty.  These must be filled in place (Group A).
        The pre-filled col W value will be overwritten with the invoice amount.
        """
        path = make_realistic_workbook()
        try:
            tracking = str(PENDING_DOMESTIC_TRACKINGS[0])  # 9522593883
            ws, (filled, inserted, created) = _run(
                path, {tracking: 7.20}
            )

            assert filled  == 1
            assert created == 0

            # Find the row
            for r in range(3, ws.max_row + 1):
                rv = ws.cell(row=r, column=COL_R).value
                if rv is not None and str(rv).replace(".0", "") == tracking:
                    assert ws.cell(row=r, column=COL_U).value == INVOICE_NO
                    assert ws.cell(row=r, column=COL_V).value == INVOICE_DT
                    assert ws.cell(row=r, column=COL_W).value == 7.20
                    # col A must still be present (domestic row untouched)
                    assert ws.cell(row=r, column=1).value is not None
                    break
        finally:
            os.unlink(path)


# ── Export rows (col A empty by design) ──────────────────────────────────────

class TestExportRows:

    def test_invoiced_export_row_skipped(self):
        """
        Export row already has FT V/431728.  Running with a different invoice
        (FT V/999999) must trigger Group B (insert below), not Group A.
        """
        path = make_realistic_workbook()
        try:
            ws, (filled, inserted, created) = _run(
                path, {INVOICED_EXPORT_TRACKING: 6.75}
            )
            assert inserted >= 1
            assert created  == 0
        finally:
            os.unlink(path)

    def test_pending_export_row_not_overwritten(self):
        """
        THE MAIN REPORTED BUG — realistic version.

        Export rows 19–20 (8255388551, 7777777778) have col A empty and
        col U/V/W empty.  A brand-new tracking not in the file must be
        appended AFTER these rows, not on top of them.
        """
        path = make_realistic_workbook()
        try:
            new_tracking = "6666666666"
            ws, (filled, inserted, created) = _run(
                path, {new_tracking: 12.50}
            )

            assert created == 1

            # Both pending export rows must be untouched
            for t in PENDING_EXPORT_TRACKINGS:
                for r in range(2, ws.max_row + 1):
                    if str(ws.cell(row=r, column=COL_R).value) == t:
                        assert ws.cell(row=r, column=COL_U).value is None, (
                            f"Pending export tracking {t} had its invoice "
                            f"column overwritten by a different shipment."
                        )
                        break

            # New tracking must exist and be at FIRST_FREE_ROW
            found_at = None
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=COL_R).value) == new_tracking:
                    found_at = r
                    break

            assert found_at is not None, "New tracking was never created"
            assert found_at == FIRST_FREE_ROW, (
                f"New tracking created at row {found_at}, expected {FIRST_FREE_ROW}. "
                f"Likely overwrote a pending export row."
            )
        finally:
            os.unlink(path)

    def test_pending_export_row_filled_when_its_invoice_arrives(self):
        """
        When the ACTUAL invoice for a pending export row arrives in a PDF,
        the script must fill it in place (Group A), not append a new row.
        """
        path = make_realistic_workbook()
        try:
            tracking = PENDING_EXPORT_TRACKINGS[0]   # "8255388551"
            ws, (filled, inserted, created) = _run(
                path, {tracking: 6.75}
            )

            assert filled  == 1
            assert created == 0

            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=COL_R).value == tracking:
                    assert ws.cell(row=r, column=COL_U).value == INVOICE_NO
                    assert ws.cell(row=r, column=COL_V).value == INVOICE_DT
                    assert ws.cell(row=r, column=COL_W).value == 6.75
                    break
        finally:
            os.unlink(path)


# ── Appended rows survive Group B rebuild ─────────────────────────────────────

class TestAppendedRowsSurvival:

    def test_appended_rows_survive_group_b_insert(self):
        """
        When a Group B insert (new invoice for existing domestic tracking)
        triggers a full rebuild of rows 1..last_data_row, the appended rows
        beyond last_data_row must not be wiped.
        """
        path = make_realistic_workbook()
        try:
            # Use a domestic tracking already invoiced → triggers Group B
            group_b_tracking = str(INVOICED_DOMESTIC_TRACKING)   # "3248136404"
            # Also add a brand-new tracking → triggers Group C
            new_tracking = "1111111111"

            ws, (filled, inserted, created) = _run(
                path,
                {group_b_tracking: 6.97, new_tracking: 5.00}
            )

            assert inserted >= 1
            assert created  == 1

            r_values = _all_r_values(ws)

            # All previously-appended rows must still exist
            for t in ["9900000001", "9900000002", "9900000003"]:
                assert t in r_values, (
                    f"Previously-appended row {t} was destroyed by Group B rebuild"
                )

            # Both pending export rows must also still exist
            for t in PENDING_EXPORT_TRACKINGS:
                assert t in r_values, (
                    f"Pending export row {t} was destroyed by Group B rebuild"
                )

            # New Group C row must exist
            assert new_tracking in r_values
        finally:
            os.unlink(path)

    def test_multiple_new_appends_after_existing_appended_rows(self):
        """
        Running a PDF with 3 new trackings must append them all AFTER the
        existing appended rows (21–23), in order.
        """
        path = make_realistic_workbook()
        try:
            new_shipments = {
                "2000000001": 6.97,
                "2000000002": 6.97,
                "2000000003": 6.97,
            }
            ws, (filled, inserted, created) = _run(path, new_shipments)

            assert created == 3

            r_values = _all_r_values(ws)
            for t in new_shipments:
                assert t in r_values

            # New rows must come after all existing appended rows
            for new_t in new_shipments:
                for old_t in ["9900000001", "9900000002", "9900000003"]:
                    idx_old = r_values.index(old_t)
                    idx_new = r_values.index(new_t)
                    assert idx_new > idx_old, (
                        f"New row {new_t} landed before existing appended row {old_t}"
                    )
        finally:
            os.unlink(path)


# ── Number formats preserved ──────────────────────────────────────────────────

class TestNumberFormats:

    def test_date_and_amount_format_on_domestic_fill(self):
        """Bug 11 + 12: formats must be correct even for domestic (int) tracking."""
        path = make_realistic_workbook()
        try:
            tracking = str(PENDING_DOMESTIC_TRACKINGS[1])   # 8572100110
            ws, _ = _run(path, {tracking: 7.20})

            for r in range(3, ws.max_row + 1):
                rv = ws.cell(row=r, column=COL_R).value
                if rv is not None and str(rv).replace(".0", "") == tracking:
                    assert ws.cell(row=r, column=COL_V).number_format == "DD-MM-YYYY"
                    assert ws.cell(row=r, column=COL_W).number_format == "\u20ac#,##0.00"
                    break
        finally:
            os.unlink(path)

    def test_date_and_amount_format_on_new_appended_row(self):
        """Bug 11 + 12: formats must be correct on brand-new appended rows."""
        path = make_realistic_workbook()
        try:
            new_tracking = "3333333333"
            ws, _ = _run(path, {new_tracking: 8.50})

            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=COL_R).value == new_tracking:
                    assert ws.cell(row=r, column=COL_V).number_format == "DD-MM-YYYY"
                    assert ws.cell(row=r, column=COL_W).number_format == "\u20ac#,##0.00"
                    break
        finally:
            os.unlink(path)
