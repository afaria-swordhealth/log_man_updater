"""
tests/fixtures/make_workbook.py
================================
Factory helpers that create realistic openpyxl workbooks for use in tests.

All helpers mirror the structure expected by update_logistics.py:
  - Sheet name  : "LogMan 2026"
  - Col A (1)   : anchor / row identifier scanned by last_data_row logic
  - Col R (18)  : tracking number  (COL_R)
  - Col U (21)  : invoice number   (COL_U)
  - Col V (22)  : invoice date     (COL_V)
  - Col W (23)  : amount EUR       (COL_W)

Header is always written on row 1; data rows start at row 2.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# ── Constants (mirror update_logistics.py) ────────────────────────────────────

SHEET_NAME = "LogMan 2026"

COL_A = 1
COL_R = 18   # tracking number
COL_U = 21   # invoice number
COL_V = 22   # invoice date
COL_W = 23   # amount EUR

# Number of columns in a full header row (A through W = 23)
MAX_COL = COL_W

# Generic header labels for columns A–W
_HEADER_LABELS = {
    1:  "Ref",
    2:  "Client",
    3:  "Origin",
    4:  "Destination",
    5:  "Weight",
    6:  "Col F",
    7:  "Col G",
    8:  "Col H",
    9:  "Col I",
    10: "Col J",
    11: "Col K",
    12: "Col L",
    13: "Col M",
    14: "Col N",
    15: "Col O",
    16: "Col P",
    17: "Col Q",
    18: "Tracking",       # R
    19: "Col S",
    20: "Col T",
    21: "Invoice No",     # U
    22: "Invoice Date",   # V
    23: "Amount EUR",     # W
}

_RED_FONT = Font(color="FFFF0000")


# ── Core factory ──────────────────────────────────────────────────────────────

def make_workbook(path, rows):
    """
    Create an openpyxl workbook at *path* that reflects the LogMan 2026 layout.

    Parameters
    ----------
    path : str | Path
        Destination file path.  The file is created (or overwritten).
    rows : list[dict]
        Each dict describes one data row (row 2 onward).  Recognised keys:

        col_a     : value for col A (the scan anchor).  None → appended row.
        tracking  : value for col R.  May be str or numeric (float/int).
        invoice   : value for col U.
        date      : datetime for col V.  Gets number_format "DD-MM-YYYY".
        amount    : float for col W.  Gets number_format "€#,##0.00".
        red       : bool.  If True, applies red font (FF0000) to U, V, W.
        hyperlink : URL string applied to col R as a hyperlink.
        merged    : bool.  If True, merges col A:B in this row (MergedCell guard).

    Returns
    -------
    Path
        Absolute path of the saved file.
    """
    path = Path(path)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # ── Row 1: header ─────────────────────────────────────────────────────────
    for col, label in _HEADER_LABELS.items():
        ws.cell(row=1, column=col, value=label)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, spec in enumerate(rows, start=2):
        col_a    = spec.get("col_a")
        tracking = spec.get("tracking")
        invoice  = spec.get("invoice")
        date     = spec.get("date")
        amount   = spec.get("amount")
        red      = spec.get("red", False)
        hyperlink = spec.get("hyperlink")
        merged   = spec.get("merged", False)

        # Col A anchor
        if col_a is not None:
            ws.cell(row=row_idx, column=COL_A, value=col_a)

        # Col R: tracking number (stored as-is; may be str or numeric)
        if tracking is not None:
            cell_r = ws.cell(row=row_idx, column=COL_R, value=tracking)
            if hyperlink:
                cell_r.hyperlink = hyperlink

        # Col U: invoice number
        if invoice is not None:
            cell_u = ws.cell(row=row_idx, column=COL_U, value=invoice)
            if red:
                cell_u.font = _RED_FONT

        # Col V: invoice date
        if date is not None:
            cell_v = ws.cell(row=row_idx, column=COL_V, value=date)
            cell_v.number_format = "DD-MM-YYYY"
            if red:
                cell_v.font = _RED_FONT

        # Col W: amount EUR
        if amount is not None:
            cell_w = ws.cell(row=row_idx, column=COL_W, value=amount)
            cell_w.number_format = "\u20ac#,##0.00"
            if red:
                cell_w.font = _RED_FONT

        # Merged cell guard: merge A:B so col A returns a MergedCell object
        if merged:
            ws.merge_cells(
                start_row=row_idx, start_column=COL_A,
                end_row=row_idx,   end_column=COL_A + 1,
            )

    wb.save(path)
    return path


# ── Convenience variants ──────────────────────────────────────────────────────

def make_clean_workbook(path, n_rows):
    """
    Create a workbook with *n_rows* normal data rows.

    Each row has:
      - col A  filled (sequential integer, e.g. 1, 2, … n)
      - col R  filled with a sequential 10-digit tracking string starting at
               "1234567890" (1234567890, 1234567891, …)
      - col U, V, W  empty  (ready to be filled by update_logistics.py)

    Useful as the baseline fixture for most unit tests.
    """
    base_tracking = 1_234_567_890
    rows = [
        {
            "col_a":    i,
            "tracking": str(base_tracking + (i - 1)),
        }
        for i in range(1, n_rows + 1)
    ]
    return make_workbook(path, rows)


def make_workbook_with_appended(path, n_main, n_appended):
    """
    Create a workbook that simulates Bug 9 (appended-row scenario).

    *n_main* rows look normal (col A present, col R filled, U/V/W empty).
    *n_appended* rows follow with col A empty and only col R filled —
    exactly the layout left by a previous run of update_logistics.py that
    used the "append at end" path.

    The extended scan ``for r in range(2, ws.max_row + 1)`` in the script
    must find the appended rows even though ``last_data_row`` stops at the
    last col-A row.
    """
    base_tracking = 1_234_567_890
    rows = []

    # Main rows (col A present)
    for i in range(1, n_main + 1):
        rows.append({
            "col_a":    i,
            "tracking": str(base_tracking + (i - 1)),
        })

    # Appended rows (col A absent — only tracking in R)
    for j in range(n_appended):
        rows.append({
            "col_a":    None,   # deliberately absent
            "tracking": str(base_tracking + n_main + j),
        })

    return make_workbook(path, rows)


def make_workbook_with_orphan_cells(path, n_rows, orphan_row):
    """
    Create a workbook with *n_rows* normal data rows plus a stray value in
    *orphan_row* in a column far outside the normal data range.

    This exercises Bug 1: a formula result or stray value in a distant column
    inflates ``ws.max_row`` / ``ws.max_column`` beyond the real data boundary.
    The ``last_data_row`` scan (col A only) must correctly ignore the orphan
    and return *n_rows* + 1 (i.e. the last real data row index).

    Parameters
    ----------
    path       : str | Path
    n_rows     : int   — number of normal data rows
    orphan_row : int   — 1-based sheet row for the orphan cell
                         (typically >> n_rows + 1, e.g. 500)
    """
    path = make_clean_workbook(path, n_rows)

    # Re-open and inject the orphan cell in a column well beyond COL_W
    from openpyxl import load_workbook as _load
    wb = _load(path)
    ws = wb[SHEET_NAME]
    orphan_col = COL_W + 10   # column 33 — outside any normal range
    ws.cell(row=orphan_row, column=orphan_col, value="ORPHAN")
    wb.save(path)
    return path
