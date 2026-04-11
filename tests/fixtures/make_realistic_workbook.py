"""
tests/fixtures/make_realistic_workbook.py
==========================================
Generates a workbook that faithfully mirrors the structure of the real
Logistics Management '26.xlsx file, without any real data.

Key structural features replicated
------------------------------------
1. Row 1  : blank (real file has a merged title row)
2. Row 2  : column headers — col 18 = "(Tracking link)", col 21 = "Invoice No.",
             col 22 = "Date", col 23 = "Amount".  The script explicitly skips
             cells whose col-R value is "(Tracking link)".
3. Domestic rows (col A filled, "OMB-XXXX")
   - Tracking stored as integer (openpyxl reads as float on reload)  ← Bug 2
   - col 23 (Amount) may be pre-filled without col 21/22 (Invoice/Date)
     (happens when a rate is known before the invoice arrives)
4. Export rows (col A empty by design)
   - col 2 = "Supply Chain & Logistics", col 3 = "Export" / "Import"
   - Tracking stored as string
5. Appended rows from a previous script run (col A empty, only R/U/V/W)
6. Orphan cell far beyond the data (inflates ws.max_row, Bug 1)
"""

import os
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font


# Column layout matching the real file
COL_A  = 1    # Ticket / order ref (OMB-XXXX for domestic; empty for export)
COL_B  = 2    # Cost Center  ("Supply Chain & Logistics")
COL_C  = 3    # Service      ("Domestic" / "Import" / "Export")
COL_D  = 4    # Flow         ("PT - PT", "CN - PT", "PT - US", …)
COL_H  = 8    # Vendor       ("DHL", "K-Log", …)
COL_I  = 9    # Shipper
COL_R  = 18   # Tracking number
COL_S  = 19   # Status       ("Delivered", "Booked", "In Transit")
COL_U  = 21   # Invoice No.
COL_V  = 22   # Invoice Date
COL_W  = 23   # Amount EUR


# ── helpers ───────────────────────────────────────────────────────────────────

_RED_FONT   = Font(color="FFFF0000")
_BLACK_FONT = Font(color="FF000000")


def _header_row(ws):
    """Write the column-header row (row 2) exactly as in the real file."""
    headers = {
        COL_A: "Ticket",
        COL_B: "Cost Center",
        COL_C: "Service",
        COL_D: "Flow",
        COL_H: "Vendor",
        COL_I: "Shipper",
        COL_R: "(Tracking link)",   # ← script skips this value in existing_rows scan
        COL_S: "Status",
        COL_U: "Invoice No.",
        COL_V: "Date",
        COL_W: "Amount",
    }
    for col, label in headers.items():
        ws.cell(row=2, column=col, value=label)


def _domestic_row(ws, row, ticket, tracking_int, status="Delivered",
                  invoice=None, inv_date=None, amount_pre=None,
                  invoice_amount=None, red_uvw=False):
    """
    Write a domestic row (col A filled).

    tracking_int  : stored as Python int so openpyxl reloads it as float  ← Bug 2
    amount_pre    : value written to col W WITHOUT a corresponding invoice
                    (mimics rows that have a rate but no invoice yet)
    invoice_amount: value for col W when invoice IS present
    """
    ws.cell(row=row, column=COL_A, value=ticket)
    ws.cell(row=row, column=COL_B, value="Supply Chain & Logistics")
    ws.cell(row=row, column=COL_C, value="Domestic")
    ws.cell(row=row, column=COL_D, value="PT - PT")
    ws.cell(row=row, column=COL_H, value="DHL")
    ws.cell(row=row, column=COL_I, value="Sword Health, SA")
    ws.cell(row=row, column=COL_R, value=tracking_int)   # int → reloads as float
    ws.cell(row=row, column=COL_S, value=status)

    if invoice:
        ws.cell(row=row, column=COL_U, value=invoice)
        ws.cell(row=row, column=COL_V, value=inv_date).number_format = "DD-MM-YYYY"
        amt_cell = ws.cell(row=row, column=COL_W, value=invoice_amount)
        amt_cell.number_format = "€#,##0.00"
        if red_uvw:
            for c in (COL_U, COL_V, COL_W):
                ws.cell(row=row, column=c).font = _RED_FONT
    elif amount_pre is not None:
        # Pre-filled amount without invoice (the W-without-U/V pattern)
        ws.cell(row=row, column=COL_W, value=amount_pre)


def _export_row(ws, row, tracking_str, flow="PT - US", vendor="K-Log",
                invoice=None, inv_date=None, amount=None):
    """
    Write an export row (col A intentionally empty — export rows have no ticket).
    Tracking stored as string.
    """
    ws.cell(row=row, column=COL_B, value="Supply Chain & Logistics")
    ws.cell(row=row, column=COL_C, value="Export")
    ws.cell(row=row, column=COL_D, value=flow)
    ws.cell(row=row, column=COL_H, value=vendor)
    ws.cell(row=row, column=COL_I, value="Sword Health, SA")
    ws.cell(row=row, column=COL_R, value=tracking_str)   # string
    ws.cell(row=row, column=COL_S, value="Booked")

    if invoice:
        ws.cell(row=row, column=COL_U, value=invoice)
        ws.cell(row=row, column=COL_V, value=inv_date).number_format = "DD-MM-YYYY"
        amt_cell = ws.cell(row=row, column=COL_W, value=amount)
        amt_cell.number_format = "€#,##0.00"


def _appended_row(ws, row, tracking_str, invoice, inv_date, amount):
    """
    Write a row that was previously appended by the script (Group C).
    Only R / U / V / W are filled; col A is empty.
    """
    ws.cell(row=row, column=COL_R, value=tracking_str)
    ws.cell(row=row, column=COL_U, value=invoice)
    ws.cell(row=row, column=COL_V, value=inv_date).number_format = "DD-MM-YYYY"
    amt_cell = ws.cell(row=row, column=COL_W, value=amount)
    amt_cell.number_format = "€#,##0.00"


def _orphan_cell(ws, row, col, formula):
    """Inject a stray formula cell far beyond the real data (Bug 1 scenario)."""
    ws.cell(row=row, column=col, value=formula)


# ── public factory ────────────────────────────────────────────────────────────

def make_realistic_workbook(path=None):
    """
    Build a workbook that mirrors the real LM '26.xlsx structure.

    Layout
    ------
    row  1  : blank
    row  2  : column headers
    rows 3-12  : domestic rows, invoice filled (tracking stored as int)
    rows 13-15 : domestic rows, col W pre-filled but U/V empty (awaiting invoice)
    rows 16-18 : export rows, col A empty, invoice filled
    rows 19-20 : export rows, col A empty, no invoice yet
    rows 21-23 : appended rows from a "previous script run" (only R/U/V/W)
    row  500   : orphan cell =sum(#REF!+#REF!) in col W  ← inflates ws.max_row

    Returns the path to the saved .xlsx file.
    """
    if path is None:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        path = tmp.name

    wb = Workbook()
    ws = wb.active
    ws.title = "LogMan 2026"

    # Row 1: blank (real file has a merged title)
    # Row 2: headers
    _header_row(ws)

    inv_a  = "FT V/395474"
    date_a = datetime(2026, 1, 12)

    # ── Domestic rows with invoice filled (rows 3–12) ────────────────────────
    domestic_invoiced = [
        (3,  "OMB-2923", 3248136404, inv_a, date_a, 6.97),
        (4,  "OMB-2924", 5335500031, inv_a, date_a, 6.97),
        (5,  "OMB-2925", 5335500032, inv_a, date_a, 6.97),
        (6,  "OMB-2926", 5335500033, inv_a, date_a, 6.97),
        (7,  "OMB-2927", 5335500034, inv_a, date_a, 6.97),
        (8,  "OMB-2928", 5335500035, inv_a, date_a, 6.97),
        (9,  "OMB-2929", 5335500036, inv_a, date_a, 6.97),
        (10, "OMB-2930", 5335500037, inv_a, date_a, 6.97),
        (11, "OMB-2931", 5335500038, inv_a, date_a, 6.97),
        (12, "OMB-2932", 5335500039, inv_a, date_a, 6.97),
    ]
    for row, ticket, tracking, inv, dt, amt in domestic_invoiced:
        _domestic_row(ws, row, ticket, tracking,
                      invoice=inv, inv_date=dt, invoice_amount=amt)

    # ── Domestic rows with W pre-filled but NO invoice (rows 13–15) ──────────
    # This is the "W-without-U/V" pattern seen in the real file (e.g. 6.97 rate
    # written before the invoice arrives).  The script WILL overwrite col W with
    # the actual invoice amount when it processes these rows.
    domestic_pending = [
        (13, "OMB-3590", 9522593883, 7.20),
        (14, "OMB-3592", 8572100110, 7.20),
        (15, "OMB-3593", 2362347956, 7.20),
    ]
    for row, ticket, tracking, pre_amt in domestic_pending:
        _domestic_row(ws, row, ticket, tracking,
                      status="Booked", amount_pre=pre_amt)

    # last_data_row = 15  (last row with col A)

    # ── Export rows with invoice filled (rows 16–18) ─────────────────────────
    export_invoiced = [
        (16, "1913526731", "FT V/431728", datetime(2026, 3, 31), 6.75),
        (17, "5001199884", "FT V/431728", datetime(2026, 3, 31), 6.75),
        (18, "3721159223", "FT V/431728", datetime(2026, 3, 31), 6.75),
    ]
    for row, tracking, inv, dt, amt in export_invoiced:
        _export_row(ws, row, tracking, invoice=inv, inv_date=dt, amount=amt)

    # ── Export rows WITHOUT invoice (rows 19–20) — col A empty, U/V/W empty ──
    # These are the critical rows: script must NOT overwrite them when
    # appending a brand-new tracking (the main reported bug).
    export_pending = [
        (19, "8255388551"),   # "Ampere" row — the exact scenario reported
        (20, "7777777778"),
    ]
    for row, tracking in export_pending:
        _export_row(ws, row, tracking)

    # last_occupied_row = 20  (last row with col R)

    # ── Appended rows from a previous script run (rows 21–23) ────────────────
    appended = [
        (21, "9900000001", "FT V/400001", datetime(2026, 2, 15), 6.97),
        (22, "9900000002", "FT V/400001", datetime(2026, 2, 15), 6.97),
        (23, "9900000003", "FT V/400001", datetime(2026, 2, 15), 6.97),
    ]
    for row, tracking, inv, dt, amt in appended:
        _appended_row(ws, row, tracking, inv, dt, amt)

    # last_occupied_row = 23

    # ── Orphan cell (Bug 1) ───────────────────────────────────────────────────
    # Mimics the real file's =sum(W899+W900+...) at row 27603 / =sum(#REF!) at 2530
    _orphan_cell(ws, 500, COL_W, "=sum(#REF!+#REF!+#REF!)")

    wb.save(path)
    return path


# ── Convenience accessors for test assertions ─────────────────────────────────

LAST_DATA_ROW     = 15   # last row with col A (OMB-3593)
LAST_OCCUPIED_ROW = 23   # last row with col R (last appended row)
FIRST_FREE_ROW    = 24   # where the next Group C append should land

# Trackings awaiting invoice (in export rows with col A empty, no U/V/W)
PENDING_EXPORT_TRACKINGS = ["8255388551", "7777777778"]

# Trackings already invoiced
INVOICED_DOMESTIC_TRACKING = 3248136404   # stored as int (Bug 2 scenario)
INVOICED_EXPORT_TRACKING   = "1913526731"

# Domestic trackings awaiting invoice (col A filled, col W pre-filled, no U/V)
PENDING_DOMESTIC_TRACKINGS = [9522593883, 8572100110, 2362347956]
